from bs4 import BeautifulSoup, Comment
import requests
from lxml import etree
from lxml import html
import os
import json
import time
from urllib.parse import urljoin
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import re
import openpyxl
from openpyxl import Workbook, styles
from openpyxl import load_workbook


# 全局变量
with open('config.json', 'r', encoding='utf-8') as f:
    JsonFile = json.load(f)
base_url = JsonFile['url']
notice_title_href_xpath = JsonFile['notice_title_href_xpath']
notice_content_xpath = JsonFile['notice_content_xpath']
notice_pages = JsonFile['notice_pages']
search = JsonFile['search']
notice_content_xpath = notice_content_xpath.replace("替换", search)
notice_title_xpath = JsonFile['notice_title_xpath']
# 全局变量不修改，调试用的
chrome_driver = r'.\chromedriver.exe'  # chromedriver的文件位置
next_page_keywords = JsonFile['next_page_keywords']
delay = JsonFile['delay']
notice_list_number = JsonFile['notice_list_number']
open_xpath_search_mode = JsonFile['open_xpath_search_mode']
content_page_number = JsonFile['content_page_number']
notice_time_xpath = JsonFile['notice_time_xpath']


# 返回下一页的链接，因为是动态js加载的，所以我要使用selenium
def get_next_page_url(current_url, notice_pages):
    s = Service(executable_path=chrome_driver)
    driver = webdriver.Chrome(service=s)
    driver.get(current_url)
    time.sleep(4)
    result = []
    result.append(driver.current_url)
    i = 1
    print(driver.current_url)
    print("开始寻找所有的公告页面")
    while True:
        try:
            if len(result) == notice_pages:
                break
            pre_url = driver.current_url
            driver.find_element(by=By.XPATH, value=next_page_keywords).click()
            if driver.current_url == pre_url:
                break
            i = i + 1
            result.append(driver.current_url)
            # if i >= notice_pages:
            #     break
            time.sleep(delay)
        except:
            print("到达页数上限")
            break
    print("已找到所有页面，一共有", i, "页")
    time.sleep(3)
    driver.quit()
    print(result)
    return result

# 自适应寻找公告列表的xpath
def get_notice_page_title(base_url):
    s = Service(executable_path=chrome_driver)
    driver = webdriver.Chrome(service=s)
    driver.get(base_url)
    time.sleep(1)
    # dynamic_html = driver.execute_script("return document.documentElement.outerHTML")
    # print(dynamic_html)
    # print(driver.execute_script("return document.documentElement.innerHTML"))
    ul_list = driver.find_elements(By.XPATH, value="//ul")
    result = ""
    # 找出最长ul
    temp = 0
    for ul in ul_list:
        # 没有字数的列表块跳过不输出
        if len(ul.text) == 0:
            continue
        elif len(ul.text) > temp:
            print("————————————————————————————————————————————")
            print("【当前最大列表块】\n字数：", len(ul.text), "\n内容：\n", ul.text)
            print("————————————————————————————————————————————")
            temp = len(ul.text)
            print(temp)
            # li_list记录的是最长板块的列表集合
            li_list = ul.find_elements(By.XPATH, value="li")
            try:
                result = li_list[notice_list_number].find_element(By.XPATH, value="a").text
            except:
                continue
            # 将列表中的第二个公告赋予xpath，防止置顶贴

        print("\n")
    print("最后得到的公告标题为：\n", result)

    return result
    # driver.quit()

# 将标题去获取页面的标题xpath
def get_xpath(title, url, search_element):
    page = requests.get(url)
    page.encoding = 'utf-8'

    # print(page.text)

    # 去除掉注释，防止getroottree获取到不正确的xpath
    soup = BeautifulSoup(page.text, 'html.parser')
    for element in soup(text=lambda text: isinstance(text, Comment)):
        element.extract()
    # print(str(soup))

    root = html.fromstring(str(soup))
    tree = root.getroottree()
    temp_xpath = "//" + search_element + "[contains(text(),\"" + title + "\")]"
    # print(temp_xpath)
    result = root.xpath(temp_xpath)
    result_xpath = tree.getpath(result[0])
    print("得到的的xpath是：\n", result_xpath)
    # 如果有多项才使用for循环
    # for r in result:
    #     print(tree.getpath(r))
    return result_xpath

# 将取得的单项xpath处理成多项，将最后的li后面中括号去掉
def handle_suffix_of_title_xpath(pre_xpath):
    # print(pre_xpath.rfind('[', beg=0, end=len(title_xpath)))
    print(str.rfind(pre_xpath, '[', 0, len(pre_xpath)))
    x = str.rfind(pre_xpath, '[', 0, len(pre_xpath))
    # print(pre_xpath)
    string_list = list(pre_xpath)
    # print(len(string_list))
    # print(string_list)
    del string_list[x:x+3]
    # print(len(string_list))
    # print(string_list)
    after_xpath = ''.join(string_list)
    # print(len(pre_xpath))
    #     # print(pre_xpath)
    #     # print(len(after_xpath))
    print("处理后的通用xpath为", after_xpath)
    return after_xpath

# 获取所有公告的链接
def get_all_notice_url_list(notice_pages_list):
    notice_list = []
    for notice_page in notice_pages_list:
        # current_url = notice_page
        # print(current_url)
        # print(notice_page)

        current_html = requests.get(notice_page)
        current_html.encoding = "utf-8"
        selector = etree.HTML(current_html.text)
        title_href_xpath = after_title_xpath + "/@href"
        notice = selector.xpath(title_href_xpath)
        # print(notice) 打印当前页面的所有公告href

        # 获取当前页面所有公告
        for result in notice:
            # 获得所有公告网址
            notice_url = urljoin(notice_page, result)
            notice_list.append(notice_url)
            # print("网址： ", notice_url)
            # # 获取文本标题
            # notice_html = requests.get(notice_url)
            # notice_html.encoding = "utf-8"
            # notice_selector = etree.HTML(notice_html.text)
            # notice_title_name_xpath = """/html/head/title/text()"""
            # notice_title_name = notice_selector.xpath(notice_title_name_xpath)
            # print("标题: ", notice_title_name[0])
    return notice_list

# 自适应获取公告的内容位置
def get_notice_content_location(all_notice_url_list):
    i = 0
    for url in all_notice_url_list:
        # 自适应加载页数
        i = i + 1
        if i > content_page_number:
            break
        # 启用selenium
        s = Service(executable_path=chrome_driver)
        driver = webdriver.Chrome(service=s)
        driver.get(url)
        time.sleep(1)
        p_list = driver.find_elements(By.XPATH, value="//p")
        # 找出字最多的p
        temp = 0
        result = ''
        for p in p_list:
            # 没有字数跳过不输出
            if len(p.text) == 0:
                continue
            elif len(p.text) > temp:
                temp = len(p.text)
                print("————————————————————————————————————————————")
                print("【当前最大块】\n字数：", len(p.text), "\n内容：\n", p.text)
                print("————————————————————————————————————————————")
                result = p.text
        print("最终获得的p字段内容为：", result)
        return result

def handle_suffix_of_content_xpath(pre_xpath):
    x = str.rfind(pre_xpath, "div", 0, len(pre_xpath))
    string_list = list(pre_xpath)
    for i in range(x, len(pre_xpath)):
        # print(string_list[i])
        if string_list[i] == '/':
            del string_list[i:len(pre_xpath)]
            break

    after_xpath = ''.join(string_list)

    print("处理后的通用xpath为\n", after_xpath)
    return after_xpath

# 【静态】将取得的通用xpath转化为内容
def get_notice(url_list, notice_content_xpath):
    result_list = []
    for url in url_list:
        print("网址： ", url)
        # 获取文本标题
        notice_html = requests.get(url)
        notice_html.encoding = "utf-8"
        notice_selector = etree.HTML(notice_html.text)
        notice_title_name = notice_selector.xpath(notice_title_xpath)[0]
        print("标题: ", notice_title_name)
        try:
            publish_time = notice_selector.xpath(notice_time_xpath)[0]
            print("时间： ", publish_time)
        except:
            publish_time = "未获取到时间"
        if not open_xpath_search_mode:
            notice_content = notice_selector.xpath(notice_content_xpath)[0].xpath("string(.)")
            # 处理多余的字符
            notice_content = notice_content.replace('\t', '').replace('\r', '')
            # 合并多余回车
            notice_content = re.sub('\n+', '\n', notice_content)
            # 去除头尾回车
            if notice_content[0] == '\n':
                notice_content = notice_content[1:]
            if notice_content[(len(notice_content)-1)] == '\n':
                notice_content = notice_content[:-1]
            print("正文：\n", "".join([s for s in notice_content.splitlines(True) if s.strip()]))
        else:
            notice_content = notice_selector.xpath(notice_content_xpath)
            notice_content = notice_content[0]
            print("正文：\n", notice_content)
        temp_dict = {}
        temp_dict.update({"url": url})
        temp_dict.update({"publish_time": publish_time})
        temp_dict.update({"title": notice_title_name})
        temp_dict.update({"content": notice_content})
        result_list.append(temp_dict)
        print(temp_dict)
    return result_list

# 创建excel文档
def create_excel():
    wb = Workbook()
    dest_filename = search + '.xlsx'
    ws1 = wb.active
    ws1.title = search
    ws1['A1'] = "时间"
    ws1['B1'] = "链接"
    ws1['C1'] = "标题"
    ws1['D1'] = "内容"
    ws2 = wb.create_sheet(title="全部")
    ws2['A1'] = "时间"
    ws2['B1'] = "链接"
    ws2['C1'] = "标题"
    ws2['D1'] = "内容"
    wb.save(filename=dest_filename)

# 保存在excel文档内
def export_to_excel(node_list):
    dest_filename = search + ".xlsx"
    wb = load_workbook(dest_filename)
    number = len(node_list)
    all_row = 2
    search_row = 2
    for node in node_list:
        publish_time = node['publish_time']
        url = node['url']
        notice_title = node['title']
        content = node['content']
        ws = wb['全部']
        ws.cell(row=all_row, column=1).value = publish_time
        ws.cell(row=all_row, column=2).value = url
        ws.cell(row=all_row, column=3).value = notice_title
        ws.cell(row=all_row, column=4).value = content
        ws.cell(row=all_row, column=4).alignment=openpyxl.styles.Alignment(wrapText=True)
        all_row = all_row + 1
        if search in content or search in notice_title:
            ws = wb[search]
            ws.cell(row=search_row, column=1).value = publish_time
            ws.cell(row=search_row, column=2).value = url
            ws.cell(row=search_row, column=3).value = notice_title
            ws.cell(row=search_row, column=4).value = content
            ws.cell(row=search_row, column=4).alignment = openpyxl.styles.Alignment(wrapText=True)
            search_row = search_row + 1
    wb.save(filename=dest_filename)
    auto_width(dest_filename, '全部')
    auto_width(dest_filename, search)

# 自动列宽
def auto_width(dest_filename, sheet_name):
    from openpyxl import load_workbook, workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(dest_filename)
    ws = wb[sheet_name]

    # 第一步：计算每列最大宽度，并存储在列表lks中。

    lks = []  # 英文变量太费劲，用汉语首字拼音代替
    for i in range(1, ws.max_column + 1):  # 每列循环
        lk = 1  # 定义初始列宽，并在每个行循环完成后重置
        for j in range(1, ws.max_row + 1):  # 每行循环
            sz = ws.cell(row=j, column=i).value  # 每个单元格内容
            if isinstance(sz, str):  # 中文占用多个字节，需要分开处理
                lk1 = len(sz.encode('gbk'))  # gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
            else:
                lk1 = len(str(sz))
            if lk < lk1:
                lk = lk1  # 借助每行循环将最大值存入lk中
            # print(lk)
        lks.append(lk)  # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）

    # 第二步：设置列宽
    for i in range(1, ws.max_column + 1):
        k = get_column_letter(i)  # 将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
        ws.column_dimensions[k].width = lks[i - 1] + 2  # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整

    wb.close()
    wb.save(dest_filename)

if __name__ == '__main__':
    if not open_xpath_search_mode:
        # 【动态】自适应获取通知列表的标题，思路是根据ul的字数最多数量判断
        title = get_notice_page_title(base_url)
        # 【静态】将上面函数获取的标题去获取页面的标题xpath
        title_xpath = get_xpath(title, base_url, "a")
        # 将取得的单项xpath处理成多项
        after_title_xpath = handle_suffix_of_title_xpath(title_xpath)
    else:
        after_title_xpath = notice_title_href_xpath

    # 【动态】自适应获取当前后续的所有页面List，防止页面使用js生成下一页链接
    notice_pages_list = get_next_page_url(base_url, notice_pages)
    # 【静态】将取得的通用xpath去获取所有的链接
    all_notice_list = get_all_notice_url_list(notice_pages_list)
    if not open_xpath_search_mode:
        # 【动态】自适应获取公告的内容位置，思路是找到字数最多的p
        notice_content_location = get_notice_content_location(all_notice_list)
        # 【静态】将上面函数获取的内容去获取页面的内容xpath
        notice_content_location_xpath = get_xpath(notice_content_location, all_notice_list[0], "p")
        # 将取得的单项xpath处理成上级div目录
        after_content_xpath = handle_suffix_of_content_xpath(notice_content_location_xpath)
    else:
        after_content_xpath = notice_content_xpath
    # 【静态】将取得的通用xpath转化为内容
    notice_list = get_notice(all_notice_list, after_content_xpath)
    # 创建excel文档
    create_excel()
    # 保存在excel文档内
    export_to_excel(notice_list)

    os.system("Pause")