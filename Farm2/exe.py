#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import sys
import json
import time
from urllib.request import urlopen
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl import load_workbook


# In[2]:


# 全局变量
with open('config.json', 'r', encoding='utf-8') as f:
    JsonFile = json.load(f)
with open('setting.json', 'r', encoding='utf-8') as f:
    SettingFile = json.load(f)
base_url = SettingFile['base_url']
wait_time = SettingFile['wait_time']
delay = SettingFile['delay']
decade_year_xpaths = SettingFile['decade_year_xpaths']
year_value_xpaths = SettingFile['year_value_xpaths']
month_value_xpaths = SettingFile['month_value_xpaths']
txt_xpaths = SettingFile['txt_xpaths']
keyword1 = SettingFile['keyword1']
keyword2 = SettingFile['keyword2']
keyword3 = SettingFile['keyword3']
keyword4 = SettingFile['keyword4']
search = JsonFile['search']
countrys = JsonFile['countrys']
mode = JsonFile['mode']
base_txt_url = SettingFile['base_txt_url']


# In[3]:


def getDriver():
    if getattr(sys, 'frozen', False):
        # 从exe包里找chromedriver依赖驱动的情况
        chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
        driver = webdriver.Chrome(chromedriver_path)
    else:
        # 普通情况下从本地文件路径找依赖的情况
        driver = webdriver.Chromedriver(executable_path=r'.\chromedriver.exe')
    return  driver


# 打开网址并且下载txt文档
def open_and_download(url):
    
    wd = driver.get(url)
    # botton_decade_year = WebDriverWait(wd, 50, 0.5).until(EC.presence_of_element_located((By.XPATH,"""//*[@id="blockMarsBranding"]/a/img""")))
    try:
        WebDriverWait(wd, wait_time, 0.5).until(
            EC.presence_of_element_located((By.XPATH,decade_year_xpaths)))
    except:
        print("开始计时")
        time.sleep(delay)
        print("结束计时")
    bottons_decade_year = driver.find_elements(By.XPATH, decade_year_xpaths)
    for botton in bottons_decade_year:
        botton.click()
    print("已点开十年")

    bottons_year_value = driver.find_elements(By.XPATH, year_value_xpaths)
    for botton in bottons_year_value:
        botton.click()
        time.sleep(0.5)
    print("已点开每年")

    bottons_month_value = driver.find_elements(By.XPATH, month_value_xpaths)
    for botton in bottons_month_value:
        try:
            botton.click()
            print("点开新月份")
            time.sleep(delay)
        except:
            time.sleep(delay)
            botton.click()
            time.sleep(delay)

    print("已点开每月")

def get_all_txt_url():
    hrefs = []
    bottons_txt = driver.find_elements(By.XPATH, txt_xpaths)
    for botton in bottons_txt:
        # print(botton.get_attribute('href'))
        hrefs.append(botton.get_attribute('href'))
    print(hrefs)
    print("已获取每份txt的URL")
    return hrefs


# In[4]:


def get_txt_content(url):
    # 处理好的文件
    content_bytes = ''
    while True:
        try:
            print("【当前url为】", url)
            text_page=urlopen(url=url)
            content_bytes = text_page.read()
            break
        except:
            print("【获取失败，等待延时】")
            time.sleep(delay)

    # print(content_bytes)
    content = str(content_bytes, encoding="utf8", errors='ignore')
    # 处理多余的字符，将他们全部转为空格
    content = content.replace('\t', ' ').replace('\r', ' ').replace('\n', ' ')
    # 将逗号句号转换为空
    content = content.replace(',', '').replace('.', '')
    # 合并多余空格
    content = re.sub(' +', ' ', content)
    print("【得到处理后的的txt内容为】", content)
    return content

def get_sheetment(content):
    temp_content = content
    # print("分割前\n", temp_content)
    if keyword1 != "":
        content_spilt = re.split(keyword1, temp_content)
        temp_content = content_spilt[1]
    # print("第一次分割\n", temp_content)
    if keyword2 != "":
        content_spilt = re.split(keyword2, temp_content)
        temp_content = content_spilt[0]
    # print("第二次分割\n", temp_content)
    print("【切分表格后的数据内容】 ", temp_content)
    return temp_content

def synonym_processing(content):
    temp_content = content
    try:
        temp_content = temp_content.replace('CHINA T', 'TAIWAN')
        temp_content = temp_content.replace('CHINA MAIN', 'CHINA')
        print("【同义词处理后】：", temp_content)
    except:
        pass
    return temp_content

def get_string_category(temp_content):
    # 返回要找的列具体值
    temp_keyword = temp_content
    # print("分割前\n", temp_keyword)
    if keyword3 != "":
        content_spilt = re.split(keyword3, temp_keyword)
        temp_keyword = content_spilt[1]
    # print("第一次分割\n", temp_keyword)
    if keyword4 != "":
        content_spilt = re.split(keyword4, temp_keyword)
        temp_keyword = content_spilt[0]
    # print("第二次分割\n", temp_keyword)
    print("【检测到的品种列表】：", temp_keyword)
    return temp_keyword

def get_int_search_category(temp_keyword, search_name):
    # 返回要查找数据的列数
    list_keyword = temp_keyword.split()
    number_location = 0
    if search_name == "TOTALS":
        number_location = len(list_keyword)
        return len(list_keyword)
    try:
        number_location=list_keyword.index(search_name)
    except:
        return -1

    print("【一共有",len(list_keyword), "种品种，所查品种在第", number_location+1, "列】")
    return number_location

def get_dict_result(temp_content, number_location, temp_url, temp_date):
    dict_result = {}
    dict_result.update({'url':temp_url})
    dict_result.update({'date':temp_date})
    if number_location == -1:
        for country in countrys:
            dict_result.update({country:0})
        return dict_result
    for country in countrys:
        print("【地区】", country, end='\t')
        numbers_content = temp_content.split(country)
        count = 0
        # 删去不包含国家的那一段
        del(numbers_content[0])
        for numbers_string in numbers_content:
            # print(numbers_string)
            num_list = re.findall('\d+', str(numbers_string))
            print(num_list[number_location],end='\t')
            count = count + int(num_list[number_location])
        print("\n【合计】", count)
        dict_result.update({country:count})
    # print(dict_result)
    return dict_result

def get_string_date(base_url):
    temp_url = base_url
    temp_url = re.split(base_txt_url, temp_url)[1]
    # print(temp_url)
    temp_url = re.split('/',temp_url)[0]
    # print(temp_url)
    print("【当前日期为】", temp_url)
    return temp_url


# In[5]:


def create_excel(search_name):
    print("【正在创建文档】")
    wb = Workbook()
    dest_filename = search_name + '.xlsx'
    ws = wb.active
    ws.title = search_name
    ws['A1'] = "Date"
    col = 2
    for country in countrys:
        ws.cell(row=1,column=col).value = country
        col = col +1
    try:
        wb.save(filename=dest_filename)
    except:
        print("【请关闭文档后重新运行】")
        os.system("Pause")


# def export_to_excel(all_result):
#     dest_filename = search + ".xlsx"
#     wb = load_workbook(dest_filename)
#     ws = wb[search]
#     all_row = 2
#     for temp_result in all_result:
#         ws.cell(row=all_row, column=1).value = temp_result['date']
#         for j in range(2, len(countrys)+2):
#             for country in countrys:
#                 str1=str(ws.cell(row=1, column=j).value)
#                 str2=str(country)
#                 if str1==str2:
#                     ws.cell(row=all_row, column=j).value = temp_result[country]
#         all_row = all_row + 1
#     wb.save(filename=dest_filename)

def export_to_temp_excel(all_result,temp_row, search_name):
    dest_filename = search_name + ".xlsx"
    wb = load_workbook(dest_filename)
    ws = wb[search_name]
    ws.cell(row=temp_row, column=1).value = all_result['date']
    for j in range(2, len(countrys)+2):
        for country in countrys:
            str1=str(ws.cell(row=1, column=j).value)
            str2=str(country)
            if str1==str2:
                ws.cell(row=temp_row, column=j).value = all_result[country]
    try:
        wb.save(filename=dest_filename)
    except:
            print("【请关闭文档后重新运行】")
            os.system("Pause")


# In[6]:


if __name__ == '__main__':
    if mode == 1:
        driver = getDriver()
        # driver = webdriver.Chrome(r'.\chromedriver.exe')
        open_and_download(base_url)
        txt_url_list = get_all_txt_url()
        save_data={'url':txt_url_list}
        with open('url.json', 'w') as obj:
            json.dump(save_data,obj)
        driver.quit()
    elif mode == 0:
        # 创建文档
        create_excel(search)
        all_row = 2
        data = json.load(open('url.json', 'r'))
        txt_url_list = data['url']
        result = []
        for url in txt_url_list:
            string_txt_all_content = get_txt_content(url)
            string_sheet_content = get_sheetment(string_txt_all_content)
            # print(string_sheet_content)
            string_after_synonyms_content = synonym_processing(string_sheet_content)
            # 返回品种列表
            string_category = get_string_category(string_after_synonyms_content)
            # 返回要查找数据的列数
            int_search_category = get_int_search_category(string_category, search)
            # if int_search_category == -1:
            #     print("【【【【【无此品种】】】】】】\n\n\n")
            #     continue
            # 返回日期
            string_date = get_string_date(url)
            # 返回国家+数量的字典
            dict_result = get_dict_result(string_after_synonyms_content, int_search_category
                                          , url, string_date)
            print(dict_result)
            # 将结果输入列表
            export_to_temp_excel(dict_result, all_row, search)
            all_row = all_row + 1
            result.append(dict_result)
            print()
    elif mode == 2:
        create_excel("CRON")
        all_row = 2
        data = json.load(open('url.json', 'r'))
        txt_url_list = data['url']
        result = []
        for url in txt_url_list:
            string_txt_all_content = get_txt_content(url)
            string_sheet_content = get_sheetment(string_txt_all_content)
            # print(string_sheet_content)
            string_after_synonyms_content = synonym_processing(string_sheet_content)
            # 返回品种列表
            string_category = get_string_category(string_after_synonyms_content)
            # 返回要查找数据的列数
            int_search_category_yellow = get_int_search_category(string_category, "YELLOW")
            int_search_category_white = get_int_search_category(string_category, "WHITE")
            # 返回日期
            string_date = get_string_date(url)
            # 返回国家+数量的字典
            dict_result_yellow = get_dict_result(string_after_synonyms_content, int_search_category_yellow
                                          , url, string_date)
            dict_result_white = get_dict_result(string_after_synonyms_content, int_search_category_white
                                          , url, string_date)
            print("yellow ", dict_result_yellow)
            print("white ",dict_result_white)
            # dict_result = merge_dict()
            dict_result = {}
            dict_result.update({'url':url})
            dict_result.update({'date':string_date})
            for country in countrys:
                count = dict_result_white[country] + dict_result_yellow[country]
                dict_result.update({country:count})
            # 将结果输入列表
            export_to_temp_excel(dict_result, all_row, "CRON")
            all_row = all_row + 1
            result.append(dict_result)
            print()


        print("【结束程序】")



